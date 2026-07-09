#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股实时买卖一体决策脚本 v8

作用：
1）读取昨晚观察池，盘中实时判断：能不能买、什么价格买、止损位在哪里。
2）读取持仓文件，盘中实时判断：继续拿、减仓、止盈、止损、是否允许加仓。
3）买点和卖点放在同一个输出里，方便你按一个脚本执行。

依赖：
  需要和 a_stock_trade_common_v7.py 放在同一个文件夹。

安装：
  python -m pip install -U akshare pandas numpy openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple

常用命令：
  # 第二天盘中看昨晚观察池
  python a_stock_live_decision_v8.py --watchlist watchlists/latest_watchlist.csv

  # 同时看观察池 + 你的持仓
  python a_stock_live_decision_v8.py --watchlist watchlists/latest_watchlist.csv --holdings holdings.csv

  # 单只股票，有持仓成本
  python a_stock_live_decision_v8.py --code 601138 --cost 66.26 --shares 100

  # 单只股票，无持仓，只看能不能买
  python a_stock_live_decision_v8.py --code 600988

免责声明：
  只做技术确认和风控提醒，不保证盈利，不构成投资建议。
"""

from __future__ import annotations

import argparse
from concurrent.futures import ProcessPoolExecutor, as_completed
import math
import time
from pathlib import Path
from datetime import datetime

import pandas as pd

from a_stock_trade_common_v7 import (
    get_hist, get_realtime_one, get_minute, score_stock, multi_factor_score, decision_from_realtime,
    get_market_context, get_sector_context, normalize_code, fmt, safe_float
)


def read_csv_smart(path: str) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        if p.name == "latest_strong_watchlist.csv":
            fallback = p.with_name("latest_watchlist.csv")
            if fallback.exists():
                print(f"找不到强买点池，暂用普通观察池：{fallback}")
                p = fallback
            else:
                raise FileNotFoundError(f"找不到文件：{path}")
        else:
            raise FileNotFoundError(f"找不到文件：{path}")
    try:
        return pd.read_csv(p, encoding="utf-8-sig")
    except Exception:
        return pd.read_csv(p, encoding="gbk")


def make_candidate_from_code(code: str, name: str = "") -> dict:
    hist, source = get_hist(code, days=360)
    try:
        r = multi_factor_score(hist, code=normalize_code(code))
    except Exception:
        r = score_stock(hist)
    factor_scores = r.get("factor_scores", {}) or {}
    return {
        "代码": normalize_code(code),
        "名称": name,
        "排名分": r["score"],
        "昨收": r["last_close"],
        "信号": r["signal"],
        "支撑1": r["support1"],
        "支撑2": r["support2"],
        "压力1": r["pressure1"],
        "压力2": r["pressure2"],
        "建议止损": r["stop"],
        "ATR14": r.get("atr14"),
        "支撑区间低": r.get("support_zone_low"),
        "支撑区间高": r.get("support_zone_high"),
        "压力区间低": r.get("pressure_zone_low"),
        "压力区间高": r.get("pressure_zone_high"),
        "箱体下沿": r.get("box_low"),
        "箱体上沿": r.get("box_high"),
        "颈线位": r.get("neckline"),
        "假突破风险": r.get("false_break_risk"),
        "关键位说明": r.get("zone_note"),
        "入选理由": "；".join(r["reasons"]),
        "主要风险": "；".join(r["risks"]),
        "factor_scores": factor_scores,
        "因子趋势": factor_scores.get("trend", ""),
        "因子动量": factor_scores.get("momentum", ""),
        "因子量价": factor_scores.get("volume", ""),
        "因子资金": factor_scores.get("flow", ""),
        "因子质量": factor_scores.get("quality", ""),
        "数据源": source,
    }


def load_holdings(path: str | None) -> dict:
    """返回 {code: holding_dict}"""
    if not path:
        return {}
    df = read_csv_smart(path)
    mp = {}
    for _, row in df.iterrows():
        code = normalize_code(row.get("代码") or row.get("code"))
        buy_date_text = str(row.get("买入日期") or row.get("buy_date") or "").strip()
        hold_days = int(safe_float(row.get("持仓天数") or row.get("hold_days"), 0))
        if hold_days <= 0 and buy_date_text and buy_date_text not in ["nan", "None", "-", "--"]:
            for fmt_text in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]:
                try:
                    buy_dt = datetime.strptime(buy_date_text, fmt_text)
                    hold_days = max(0, (datetime.now() - buy_dt).days)
                    break
                except Exception:
                    pass
        mp[code] = {
            "代码": code,
            "名称": row.get("名称") or row.get("name") or "",
            "成本": safe_float(row.get("成本") or row.get("cost")),
            "股数": int(safe_float(row.get("股数") or row.get("shares"), 0)),
            "持仓天数": hold_days,
            "买入日期": buy_date_text,
        }
    return mp


def merge_watchlist_and_holdings(watchlist_path: str | None, holdings_path: str | None, code: str | None, cost: float | None, shares: int | None) -> list[dict]:
    """合并观察池和持仓。持仓优先标记为已有仓位。"""
    items = {}

    # 观察池
    if watchlist_path:
        df = read_csv_smart(watchlist_path)
        for _, row in df.iterrows():
            d = row.to_dict()
            c = normalize_code(d.get("代码") or d.get("code"))
            d["代码"] = c
            d["是否持仓"] = False
            items[c] = d

    # 持仓
    holdings = load_holdings(holdings_path)
    for c, h in holdings.items():
        if c in items:
            items[c].update({
                "是否持仓": True,
                "成本": h["成本"],
                "股数": h["股数"],
                "持仓天数": h.get("持仓天数", 0),
                "买入日期": h.get("买入日期", ""),
                "名称": items[c].get("名称") or h.get("名称", ""),
            })
        else:
            # 持仓不一定在昨晚观察池，单独计算日线候选信息
            cand = make_candidate_from_code(c, h.get("名称", ""))
            cand.update({
                "是否持仓": True,
                "成本": h["成本"],
                "股数": h["股数"],
                "持仓天数": h.get("持仓天数", 0),
                "买入日期": h.get("买入日期", ""),
            })
            items[c] = cand

    # 单只股票
    if code:
        c = normalize_code(code)
        cand = make_candidate_from_code(c)
        cand["是否持仓"] = bool(cost and shares)
        if cost:
            cand["成本"] = cost
        if shares:
            cand["股数"] = shares
        items[c] = cand

    return list(items.values())


def _failed_check_text(checks: list[tuple[str, bool]], limit: int = 4) -> str:
    failed = [name for name, ok in checks if not ok]
    return "；".join(failed[:limit])


def _fmt_range(low: float, high: float) -> str:
    if math.isnan(low) or math.isnan(high):
        return "-"
    return f"{fmt(low)} ~ {fmt(high)}"


def _round_lot(shares: float) -> int:
    if math.isnan(safe_float(shares)) or shares <= 0:
        return 0
    return int(shares // 100 * 100)


def _build_outlook(d: dict) -> str:
    parts = []
    market = str(d.get("市场环境") or "")
    sector = str(d.get("板块强弱") or "")
    relative = str(d.get("相对强弱") or "")
    level_state = str(d.get("位阶状态") or "")
    if market == "弱势":
        parts.append("大盘偏弱，反弹先按修复看")
    elif market == "强势":
        parts.append("大盘环境支持小仓试错")
    elif market == "震荡":
        parts.append("大盘分化，只看板块和位置")
    if sector == "强":
        parts.append("板块强，顺势票胜率更高")
    elif sector == "弱":
        parts.append("板块弱，冲高容易反复")
    if relative == "强于板块":
        parts.append("个股强于板块")
    elif relative == "弱于板块":
        parts.append("个股弱于板块")
    if "突破" in level_state:
        parts.append("突破后以回踩不破确认有效")
    elif "压力" in level_state:
        parts.append("临近压力，先看能否放量站上")
    elif "跌破" in level_state:
        parts.append("结构转弱，先按风控处理")
    else:
        parts.append("按支撑压力区间高抛低吸")
    return "；".join(parts)


def get_live_decision(
    item: dict,
    minute_period: str = "1",
    market_context: dict | None = None,
    capital: float | None = None,
    account_risk_pct: float = 1.0,
) -> dict:
    code = normalize_code(item.get("代码") or item.get("code"))
    name = item.get("名称", "")
    is_holding = bool(item.get("是否持仓", False))
    cost = safe_float(item.get("成本"), float("nan"))
    shares = int(safe_float(item.get("股数"), 0)) if is_holding else 0

    rt = get_realtime_one(code)
    minute = get_minute(code, period=minute_period)
    sector_context = get_sector_context(code, name=name, spot_row=rt or {}, candidate=item)
    market_context = market_context or get_market_context()

    # 同时算买入确认和卖出/持仓确认
    buy_d = decision_from_realtime(item, rt, minute, mode="buy", sector_context=sector_context, market_context=market_context)
    sell_d = decision_from_realtime(item, rt, minute, mode="sell", cost=cost if is_holding else None, shares=shares if is_holding else None, sector_context=sector_context, market_context=market_context)

    price = safe_float(buy_d.get("当前价"))
    support = safe_float(buy_d.get("当前有效支撑"))
    pressure = safe_float(buy_d.get("当前有效压力"))
    defense = safe_float(buy_d.get("防守位"))
    stop = safe_float(buy_d.get("止损位"))
    buy_low = safe_float(buy_d.get("买入区间低"))
    buy_high = safe_float(buy_d.get("买入区间高"))
    pressure_state = str(buy_d.get("压力状态") or "")
    level_state = str(buy_d.get("位阶状态") or "")
    data_quality = safe_float(buy_d.get("数据质量分"), float("nan"))
    data_gap = str(buy_d.get("数据缺口") or "")
    stop_risk_pct = safe_float(buy_d.get("止损距离%"))
    target_1r = safe_float(buy_d.get("1R目标价"))
    target_2r = safe_float(buy_d.get("2R目标价"))
    atr_stop = safe_float(buy_d.get("ATR止损位"))
    atr14 = safe_float(buy_d.get("ATR14"))
    breakout_text = str(buy_d.get("已突破位") or "")
    reclaim_text = str(buy_d.get("待收复位") or "")
    invalid_line = str(buy_d.get("失效条件") or "")
    position_suggestion = str((sell_d if is_holding else buy_d).get("仓位建议") or "")
    review_note = str(buy_d.get("复核提示") or "")
    buy_zone_valid = bool(buy_d.get("买入区间有效", False))
    breakout_pressure = bool(buy_d.get("突破昨晚压力", False))

    pnl_pct = float("nan")
    pnl_money = float("nan")
    if is_holding and not math.isnan(cost) and cost > 0 and shares:
        pnl_pct = (price - cost) / cost * 100
        pnl_money = (price - cost) * shares

    # 统一最终动作
    risk_notes = buy_d.get("风险提示", [])
    sell_action = sell_d.get("持仓建议", "未持仓")
    buy_action = buy_d.get("买入建议", "不建议买")
    pass_count = buy_d.get("通过数", 0)
    check_total = buy_d.get("检查总数", 0)
    factor_trend = safe_float(buy_d.get("因子趋势"), float("nan"))
    factor_momentum = safe_float(buy_d.get("因子动量"), float("nan"))
    factor_volume = safe_float(buy_d.get("因子量价"), float("nan"))
    factor_flow = safe_float(buy_d.get("因子资金"), float("nan"))
    factor_quality = safe_float(buy_d.get("因子质量"), float("nan"))

    if is_holding:
        if "止损" in sell_action or "减仓" in sell_action:
            final_action = sell_action
        elif "止盈" in sell_action:
            final_action = sell_action
        elif buy_action == "可以买小仓" and pass_count >= 5 and (math.isnan(pnl_pct) or pnl_pct > -2):
            final_action = "可继续持有；激进者可小幅加仓"
        else:
            final_action = "继续持有观察"
    else:
        final_action = buy_action
        if buy_action == "可以买小仓" and str(buy_d.get("市场环境") or "") == "弱势":
            final_action = "弱势市场仅可3%试错仓"

    no_buy_reason = ""
    if buy_action != "可以买小仓":
        no_buy_reason = "；".join([x for x in risk_notes if x])
        failed_text = _failed_check_text(buy_d.get("检查项", []))
        if failed_text:
            no_buy_reason = "；".join([x for x in [no_buy_reason, failed_text] if x])

    # 更清楚的价格计划
    buy_plan = "不买"
    if not is_holding:
        if buy_action == "可以买小仓":
            if not buy_zone_valid:
                buy_plan = "买入区间无效，先不下单"
            elif breakout_pressure:
                buy_plan = f"已突破旧压力 {breakout_text or fmt(support)}；只等回踩 {fmt(buy_low)} ~ {fmt(buy_high)} 不破小仓，不追高"
            else:
                buy_plan = f"可在 {fmt(buy_low)} ~ {fmt(buy_high)} 小仓试买；跌破 {fmt(stop)} 止损"
        elif buy_action == "谨慎观察，等回踩确认":
            if breakout_pressure:
                buy_plan = f"已突破旧压力 {breakout_text or fmt(support)}，但当前不适合追；等回踩突破位附近确认"
            elif reclaim_text:
                buy_plan = f"上方有待收复位 {reclaim_text}，先看能否站回；不提前买"
            else:
                buy_plan = f"等回踩 {fmt(support)} 附近不破再看；不要追高"
        else:
            if breakout_pressure:
                buy_plan = f"已突破旧压力 {breakout_text or fmt(support)}，但当前涨幅/风险不适合新买；等回踩确认"
            elif reclaim_text:
                buy_plan = f"尚未收复 {reclaim_text}，不建议新买"
            else:
                buy_plan = "不建议新买"
    else:
        if buy_action == "可以买小仓" and pass_count >= 5:
            if breakout_pressure:
                buy_plan = f"已有仓位可持有；加仓只等回踩突破位 {breakout_text or fmt(support)} 附近确认，不追高"
            else:
                buy_plan = f"已有仓位可持有；加仓只考虑 {fmt(buy_low)} ~ {fmt(buy_high)} 小幅加，不可满仓"
        else:
            buy_plan = "已有仓位不建议加仓"

    sell_plan = "未持仓，只设置预案"
    if is_holding:
        if not math.isnan(stop) and price <= stop:
            sell_plan = f"已触及/接近止损 {fmt(stop)}，优先减仓或止损"
        elif "减仓" in sell_action:
            reason = "；".join(sell_d.get("卖出理由", []))
            sell_plan = reason or f"跌破有效防守位 {fmt(defense)}，建议减仓防守"
        elif breakout_pressure:
            sell_plan = f"已突破旧压力 {breakout_text or fmt(support)}；持仓以防守位 {fmt(defense)} 跟踪，跌回去再减仓"
        elif not math.isnan(pressure) and price >= pressure * 0.98:
            sell_plan = f"接近压力 {fmt(pressure)}，建议分批止盈/减仓"
        elif not math.isnan(pnl_pct) and pnl_pct <= -5:
            sell_plan = f"亏损 {fmt(pnl_pct)}%，达到纪律止损区，建议减仓"
        elif not math.isnan(pnl_pct) and pnl_pct >= 10:
            sell_plan = f"盈利 {fmt(pnl_pct)}%，建议至少止盈一部分"
        else:
            sell_plan = f"暂未触发卖点；跌破防守位 {fmt(defense)} 或止损 {fmt(stop)} 再处理"
    else:
        if breakout_pressure:
            sell_plan = f"若后续回踩确认后买入，突破位 {breakout_text or fmt(support)} 可作为防守参考；止损看 {fmt(stop)}"
        else:
            sell_plan = f"若买入，止损看 {fmt(stop)}；第一压力/止盈参考 {fmt(pressure)}"

    if is_holding:
        if "止损" in final_action or "减仓" in final_action or "止盈" in final_action:
            now_action = final_action
        elif buy_action == "可以买小仓" and pass_count >= 5 and (math.isnan(pnl_pct) or pnl_pct > -2):
            now_action = "继续持有；加仓只等回踩确认"
        else:
            now_action = "继续持有观察；暂不加仓"
        can_buy_text = "已有仓位，按加仓条件处理" if buy_action == "可以买小仓" else "已有仓位，不建议加仓"
    else:
        if buy_action == "可以买小仓":
            now_action = "可以买小仓"
            can_buy_text = "可以买小仓"
        elif buy_action == "谨慎观察，等回踩确认":
            now_action = "不追，等回踩确认"
            can_buy_text = "暂不买，等回踩确认"
        else:
            now_action = "不买"
            can_buy_text = "不能买"

    suggested_buy_price = _fmt_range(buy_low, buy_high) if buy_action == "可以买小仓" else "不买"
    if buy_action == "谨慎观察，等回踩确认" and not math.isnan(support):
        suggested_buy_price = f"等 {fmt(support)} 附近回踩不破"
    if is_holding and buy_action != "可以买小仓":
        suggested_buy_price = "不加仓"
    first_take_profit = pressure
    if not math.isnan(target_1r):
        first_take_profit = min(first_take_profit, target_1r) if not math.isnan(first_take_profit) else target_1r
    strong_pressure = safe_float(buy_d.get("压力区间高"), safe_float(buy_d.get("压力2")))
    sell_price_plan = f"跌破 {fmt(defense)} 防守；跌破 {fmt(stop)} 止损；第一止盈/减仓看 {fmt(first_take_profit)}"
    if not math.isnan(target_2r):
        sell_price_plan += f"；2R目标 {fmt(target_2r)}"
    if not math.isnan(strong_pressure):
        sell_price_plan += f"；强压力看 {fmt(strong_pressure)}"

    suggested_shares = 0
    suggested_amount = float("nan")
    risk_budget = float("nan")
    if capital and capital > 0 and buy_action == "可以买小仓" and not math.isnan(buy_low) and not math.isnan(stop) and buy_low > stop:
        risk_budget = capital * max(account_risk_pct, 0.1) / 100
        if not math.isnan(stop_risk_pct) and stop_risk_pct > 6.5:
            risk_budget *= 0.5
        market_state = str(buy_d.get("市场环境") or "")
        if market_state == "弱势":
            risk_budget *= 0.4
        elif market_state == "震荡":
            risk_budget *= 0.7
        suggested_shares = _round_lot(risk_budget / max(buy_low - stop, 0.01))
        suggested_amount = suggested_shares * buy_low if suggested_shares else float("nan")
    position_calc_text = ""
    if capital and capital > 0:
        if suggested_shares:
            position_calc_text = f"按账户风险 {fmt(account_risk_pct)}% 估算，建议约 {suggested_shares} 股，金额约 {fmt(suggested_amount)} 元"
        elif buy_action == "可以买小仓":
            position_calc_text = "按当前止损距离估算不足100股，建议放弃或降低标的价格/风险"
        else:
            position_calc_text = "当前不满足买入条件，不计算买入股数"

    return {
        "代码": code,
        "名称": name,
        "是否持仓": is_holding,
        "成本": cost,
        "股数": shares,
        "当前价": price,
        "涨跌幅": safe_float(buy_d.get("涨跌幅")),
        "通过数": pass_count,
        "检查总数": check_total,
        "数据质量分": data_quality,
        "数据缺口": data_gap,
        "昨晚信号": buy_d.get("昨晚信号", ""),
        "买入判断": buy_action,
        "因子趋势": factor_trend,
        "因子动量": factor_momentum,
        "因子量价": factor_volume,
        "因子资金": factor_flow,
        "因子质量": factor_quality,
        "压力状态": pressure_state,
        "位阶状态": level_state,
        "当前有效支撑": support,
        "当前有效压力": pressure,
        "已突破位": breakout_text,
        "待收复位": reclaim_text,
        "防守位": defense,
        "失效条件": invalid_line,
        "仓位建议": position_suggestion,
        "止损距离%": stop_risk_pct,
        "复核提示": review_note,
        "买入计划": buy_plan,
        "卖出计划": sell_plan,
        "最终动作": final_action,
        "现在动作": now_action,
        "能不能买": can_buy_text,
        "建议买入价": suggested_buy_price,
        "建议卖出价": sell_price_plan,
        "第一止盈价": first_take_profit,
        "强压力价": strong_pressure,
        "1R目标价": target_1r,
        "2R目标价": target_2r,
        "ATR14": atr14,
        "ATR止损位": atr_stop,
        "防守价": defense,
        "建议股数": suggested_shares,
        "建议金额": suggested_amount,
        "账户风险预算": risk_budget,
        "仓位计算": position_calc_text,
        "不买原因": no_buy_reason,
        "后续行情展望": _build_outlook({
            "市场环境": buy_d.get("市场环境", ""),
            "板块强弱": buy_d.get("板块强弱", ""),
            "相对强弱": buy_d.get("相对强弱", ""),
            "位阶状态": level_state,
        }),
        "市场环境": buy_d.get("市场环境", ""),
        "市场建议": buy_d.get("市场建议", ""),
        "所属板块": buy_d.get("所属板块", ""),
        "板块涨跌幅": safe_float(buy_d.get("板块涨跌幅")),
        "板块3日涨跌幅": safe_float(buy_d.get("板块3日涨跌幅")),
        "板块5日涨跌幅": safe_float(buy_d.get("板块5日涨跌幅")),
        "板块持续性": buy_d.get("板块持续性", ""),
        "板块排名": buy_d.get("板块排名", ""),
        "板块总数": buy_d.get("板块总数", ""),
        "板块强弱": buy_d.get("板块强弱", ""),
        "板块提示": buy_d.get("板块提示", ""),
        "个股相对板块": safe_float(buy_d.get("个股相对板块")),
        "相对强弱": buy_d.get("相对强弱", ""),
        "支撑区间低": safe_float(buy_d.get("支撑区间低")),
        "支撑区间高": safe_float(buy_d.get("支撑区间高")),
        "压力区间低": safe_float(buy_d.get("压力区间低")),
        "压力区间高": safe_float(buy_d.get("压力区间高")),
        "箱体下沿": safe_float(buy_d.get("箱体下沿")),
        "箱体上沿": safe_float(buy_d.get("箱体上沿")),
        "颈线位": safe_float(buy_d.get("颈线位")),
        "关键位说明": buy_d.get("关键位说明", ""),
        "支撑1": safe_float(buy_d.get("支撑1")),
        "支撑2": safe_float(buy_d.get("支撑2")),
        "压力1": safe_float(buy_d.get("压力1")),
        "压力2": safe_float(buy_d.get("压力2")),
        "止损位": stop,
        "买入区间低": buy_low,
        "买入区间高": buy_high,
        "盈亏比例": pnl_pct,
        "盈亏金额": pnl_money,
        "风险提示": "；".join(risk_notes),
        "买入检查项": buy_d.get("检查项", []),
        "卖出理由": "；".join(sell_d.get("卖出理由", [])),
    }


def print_decision(d: dict, show_checks: bool = False):
    print("=" * 100)
    holding_text = "持仓" if d["是否持仓"] else "未持仓"
    check_total = d.get("检查总数") or len(d.get("买入检查项", [])) or 0
    print(f"{d['代码']} {d.get('名称','')} | {holding_text} | 当前价 {fmt(d['当前价'])} | 涨跌幅 {fmt(d['涨跌幅'])}% | 通过 {d['通过数']}/{check_total}")
    print(f"昨晚信号：{d.get('昨晚信号','')}")
    if d["是否持仓"]:
        print(f"持仓：成本 {fmt(d['成本'])} | 股数 {d['股数']} | 盈亏 {fmt(d['盈亏比例'])}% | 盈亏金额 {fmt(d['盈亏金额'])} 元")
    print(f"现在动作：{d.get('现在动作', d['最终动作'])}")
    print(f"能不能买：{d.get('能不能买', d['买入判断'])} | 建议买入价：{d.get('建议买入价', '-')}")
    if d.get("仓位计算"):
        print(f"仓位计算：{d['仓位计算']}")
    print(f"建议卖出价：{d.get('建议卖出价', '-')}")
    print(f"最终动作：{d['最终动作']}")
    print(f"买入判断：{d['买入判断']}")
    if any(not math.isnan(safe_float(d.get(k), float("nan"))) for k in ["因子趋势", "因子动量", "因子量价", "因子资金", "因子质量"]):
        print(f"因子：趋势 {fmt(d.get('因子趋势'))} | 动量 {fmt(d.get('因子动量'))} | 量价 {fmt(d.get('因子量价'))} | 资金 {fmt(d.get('因子资金'))} | 质量 {fmt(d.get('因子质量'))}")
    if d.get("市场环境"):
        print(f"大盘环境：{d.get('市场环境')} | {d.get('市场建议', '')}")
    if d.get("数据质量分") != "":
        print(f"数据质量：{fmt(d.get('数据质量分'), 0)} | 缺口：{d.get('数据缺口') or '无'}")
    if d.get("所属板块"):
        rank_text = f"{d.get('板块排名')}/{d.get('板块总数')}" if d.get("板块排名") else "-"
        print(f"板块：{d.get('所属板块')} | {d.get('板块强弱')} | 涨跌幅 {fmt(d.get('板块涨跌幅'))}% | 排名 {rank_text} | 个股相对板块 {fmt(d.get('个股相对板块'))}%")
        if d.get("板块持续性"):
            print(f"板块持续性：{d.get('板块持续性')} | 3日 {fmt(d.get('板块3日涨跌幅'))}% | 5日 {fmt(d.get('板块5日涨跌幅'))}%")
    if d.get("位阶状态"):
        print(f"位阶状态：{d['位阶状态']}")
    if d.get("压力状态"):
        print(f"压力状态：{d['压力状态']}")
    if d.get("关键位说明"):
        print(f"关键位：{d['关键位说明']}")
    print(f"买入计划：{d['买入计划']}")
    print(f"卖出计划：{d['卖出计划']}")
    print(f"当前有效支撑：{fmt(d.get('当前有效支撑'))} | 当前有效压力：{fmt(d.get('当前有效压力'))} | 防守位：{fmt(d.get('防守位'))} | 止损：{fmt(d['止损位'])}")
    print(f"第一止盈：{fmt(d.get('第一止盈价'))} | 1R：{fmt(d.get('1R目标价'))} | 2R：{fmt(d.get('2R目标价'))} | 强压力：{fmt(d.get('强压力价'))}")
    if not math.isnan(safe_float(d.get("ATR14"))):
        print(f"ATR14：{fmt(d.get('ATR14'))} | ATR止损：{fmt(d.get('ATR止损位'))}")
    if d.get("已突破位") or d.get("待收复位"):
        print(f"已突破位：{d.get('已突破位') or '-'} | 待收复位：{d.get('待收复位') or '-'}")
    if d.get("失效条件"):
        print(f"失效条件：{d['失效条件']}")
    if d.get("后续行情展望"):
        print(f"后续展望：{d['后续行情展望']}")
    if d.get("不买原因"):
        print(f"不买原因：{d['不买原因']}")
    if d.get("风险提示"):
        print(f"风险提示：{d['风险提示']}")
    if d.get("复核提示"):
        print(f"复核提示：{d['复核提示']}")
    if d.get("卖出理由"):
        print(f"卖出理由：{d['卖出理由']}")
    if show_checks:
        print("买入检查项：")
        for name, ok in d.get("买入检查项", []):
            print(f"  {'√' if ok else '×'} {name}")


def action_tag(r: dict) -> str:
    if not r.get("是否持仓"):
        if r.get("买入判断") == "可以买小仓":
            return "可买入"
        text = f"{r.get('最终动作', '')} {r.get('买入判断', '')} {r.get('现在动作', '')}"
        if "观察" in text:
            return "观察"
        return "不买/无动作"

    text = f"{r.get('最终动作', '')} {r.get('现在动作', '')} {r.get('卖出理由', '')}"
    if "止损" in text:
        return "止损/风控"
    if "减仓" in text:
        return "减仓"
    if "止盈" in text:
        return "止盈"
    if "加仓" in text:
        return "可加仓"
    if r.get("买入判断") == "可以买小仓":
        return "可买入"
    if "观察" in text:
        return "观察"
    return "不买/无动作"


def action_priority(tag: str) -> int:
    order = {
        "止损/风控": 1,
        "减仓": 2,
        "止盈": 3,
        "可买入": 4,
        "可加仓": 5,
        "观察": 6,
        "不买/无动作": 7,
    }
    return order.get(tag, 99)


def write_excel_report(df: pd.DataFrame, xlsx_path: Path) -> None:
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="实时决策")
        summary = (
            df.groupby("操作类型", dropna=False)
            .size()
            .reset_index(name="数量")
            .sort_values("操作类型")
        )
        summary.to_excel(writer, index=False, sheet_name="汇总")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(xlsx_path)
    ws = wb["实时决策"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fills = {
        "止损/风控": "F4CCCC",
        "减仓": "FCE5CD",
        "止盈": "FFF2CC",
        "可买入": "D9EAD3",
        "可加仓": "D0E0E3",
        "观察": "D9EAF7",
    }
    for row in range(2, ws.max_row + 1):
        tag = ws.cell(row=row, column=2).value
        fill_color = fills.get(tag)
        if fill_color:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_color)

    pct_cols = {"涨跌幅", "盈亏比例", "板块涨跌幅", "板块3日涨跌幅", "板块5日涨跌幅", "个股相对板块", "止损距离%"}
    money_cols = {
        "成本", "当前价", "买入区间低", "买入区间高", "支撑1", "支撑2", "压力1", "压力2",
        "当前有效支撑", "当前有效压力", "防守位", "防守价", "止损位", "盈亏金额",
        "第一止盈价", "强压力价", "支撑区间低", "支撑区间高", "压力区间低", "压力区间高",
        "箱体下沿", "箱体上沿", "颈线位", "1R目标价", "2R目标价", "ATR14", "ATR止损位",
        "建议金额", "账户风险预算",
    }
    for idx, cell in enumerate(ws[1], start=1):
        name = cell.value
        width = 12
        if name in ["最终动作", "现在动作", "买入计划", "卖出计划", "风险提示", "卖出理由", "失效条件", "复核提示", "建议卖出价", "不买原因", "后续行情展望", "关键位说明", "市场建议", "板块提示", "数据缺口", "仓位计算"]:
            width = 34
        elif name in ["名称", "昨晚信号", "买入判断", "能不能买", "建议买入价", "位阶状态", "压力状态", "仓位建议", "所属板块", "板块强弱", "相对强弱", "市场环境"]:
            width = 18
        elif name in ["时间"]:
            width = 20
        ws.column_dimensions[cell.column_letter].width = width
        if name == "代码":
            for row in range(2, ws.max_row + 1):
                code_cell = ws.cell(row=row, column=idx)
                code_cell.value = str(code_cell.value).split(".")[0].zfill(6)
                code_cell.number_format = "@"
        if name in pct_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "0.00"
        if name in money_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "0.00"

    if "盈亏比例" in [c.value for c in ws[1]]:
        pnl_col = [c.value for c in ws[1]].index("盈亏比例") + 1
        letter = ws.cell(row=1, column=pnl_col).column_letter
        rng = f"{letter}2:{letter}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=Font(color="9C0006")))
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="006100")))

    for sheet_name in ["实时决策", "汇总"]:
        ws2 = wb[sheet_name]
        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(xlsx_path)


def round_or_blank(value, digits: int = 3):
    v = safe_float(value)
    if math.isnan(v):
        return ""
    return round(v, digits)


def save_results(rows: list[dict]):
    if not rows:
        return
    Path("live_reports").mkdir(exist_ok=True)
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df = pd.DataFrame([{
        "时间": now_text,
        "操作类型": action_tag(r),
        "代码": str(r["代码"]).zfill(6),
        "名称": r["名称"],
        "是否持仓": "是" if r["是否持仓"] else "否",
        "成本": round_or_blank(r["成本"]),
        "股数": r["股数"],
        "当前价": round_or_blank(r["当前价"]),
        "涨跌幅": round_or_blank(r["涨跌幅"], 2),
        "通过数": f"{r.get('通过数', 0)}/{r.get('检查总数', '')}",
        "_通过数值": r.get("通过数", 0),
        "数据质量分": round_or_blank(r.get("数据质量分"), 0),
        "数据缺口": r.get("数据缺口", ""),
        "现在动作": r.get("现在动作", ""),
        "能不能买": r.get("能不能买", ""),
        "建议买入价": r.get("建议买入价", ""),
        "建议卖出价": r.get("建议卖出价", ""),
        "建议股数": r.get("建议股数", ""),
        "建议金额": round_or_blank(r.get("建议金额"), 2),
        "账户风险预算": round_or_blank(r.get("账户风险预算"), 2),
        "仓位计算": r.get("仓位计算", ""),
        "买入判断": r["买入判断"],
        "因子趋势": round_or_blank(r.get("因子趋势"), 2),
        "因子动量": round_or_blank(r.get("因子动量"), 2),
        "因子量价": round_or_blank(r.get("因子量价"), 2),
        "因子资金": round_or_blank(r.get("因子资金"), 2),
        "因子质量": round_or_blank(r.get("因子质量"), 2),
        "最终动作": r["最终动作"],
        "市场环境": r.get("市场环境", ""),
        "市场建议": r.get("市场建议", ""),
        "所属板块": r.get("所属板块", ""),
        "板块强弱": r.get("板块强弱", ""),
        "板块涨跌幅": round_or_blank(r.get("板块涨跌幅"), 2),
        "板块3日涨跌幅": round_or_blank(r.get("板块3日涨跌幅"), 2),
        "板块5日涨跌幅": round_or_blank(r.get("板块5日涨跌幅"), 2),
        "板块持续性": r.get("板块持续性", ""),
        "板块排名": r.get("板块排名", ""),
        "板块总数": r.get("板块总数", ""),
        "个股相对板块": round_or_blank(r.get("个股相对板块"), 2),
        "相对强弱": r.get("相对强弱", ""),
        "位阶状态": r.get("位阶状态", ""),
        "压力状态": r.get("压力状态", ""),
        "当前有效支撑": round_or_blank(r.get("当前有效支撑")),
        "当前有效压力": round_or_blank(r.get("当前有效压力")),
        "已突破位": r.get("已突破位", ""),
        "待收复位": r.get("待收复位", ""),
        "防守位": round_or_blank(r.get("防守位")),
        "失效条件": r.get("失效条件", ""),
        "仓位建议": r.get("仓位建议", ""),
        "止损距离%": round_or_blank(r.get("止损距离%"), 2),
        "第一止盈价": round_or_blank(r.get("第一止盈价")),
        "1R目标价": round_or_blank(r.get("1R目标价")),
        "2R目标价": round_or_blank(r.get("2R目标价")),
        "ATR14": round_or_blank(r.get("ATR14")),
        "ATR止损位": round_or_blank(r.get("ATR止损位")),
        "强压力价": round_or_blank(r.get("强压力价")),
        "防守价": round_or_blank(r.get("防守价")),
        "买入区间低": round_or_blank(r["买入区间低"]),
        "买入区间高": round_or_blank(r["买入区间高"]),
        "支撑区间低": round_or_blank(r.get("支撑区间低")),
        "支撑区间高": round_or_blank(r.get("支撑区间高")),
        "压力区间低": round_or_blank(r.get("压力区间低")),
        "压力区间高": round_or_blank(r.get("压力区间高")),
        "箱体下沿": round_or_blank(r.get("箱体下沿")),
        "箱体上沿": round_or_blank(r.get("箱体上沿")),
        "颈线位": round_or_blank(r.get("颈线位")),
        "关键位说明": r.get("关键位说明", ""),
        "支撑1": round_or_blank(r["支撑1"]),
        "支撑2": round_or_blank(r.get("支撑2")),
        "压力1": round_or_blank(r["压力1"]),
        "压力2": round_or_blank(r.get("压力2")),
        "止损位": round_or_blank(r["止损位"]),
        "盈亏比例": round_or_blank(r["盈亏比例"], 2),
        "盈亏金额": round_or_blank(r["盈亏金额"], 2),
        "风险提示": r["风险提示"],
        "不买原因": r.get("不买原因", ""),
        "后续行情展望": r.get("后续行情展望", ""),
        "板块提示": r.get("板块提示", ""),
        "复核提示": r.get("复核提示", ""),
        "卖出理由": r["卖出理由"],
        "买入计划": r["买入计划"],
        "卖出计划": r["卖出计划"],
    } for r in rows])
    df["_priority"] = df["操作类型"].apply(action_priority)
    df = df.sort_values(["_priority", "是否持仓", "_通过数值"], ascending=[True, False, False]).drop(columns=["_priority", "_通过数值"])

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = Path("live_reports") / f"live_decision_{stamp}.csv"
    latest_csv = Path("live_reports") / "latest_live_decision.csv"
    xlsx_path = Path("live_reports") / f"live_decision_{stamp}.xlsx"
    latest_xlsx = Path("live_reports") / "latest_live_decision.xlsx"

    saved_files = []
    for path in [csv_path, latest_csv]:
        try:
            df.to_csv(path, index=False, encoding="utf-8-sig")
            saved_files.append(path)
        except PermissionError:
            print(f"\n文件被占用，已跳过覆盖：{path}")
        except Exception as e:
            print(f"\nCSV 保存失败：{path}：{e}")

    for path in [xlsx_path, latest_xlsx]:
        try:
            write_excel_report(df, path)
            saved_files.append(path)
        except PermissionError:
            print(f"\n文件被占用，已跳过覆盖：{path}")
        except Exception as e:
            print(f"\nExcel 保存失败：{path}：{e}")

    print("\n已保存实时买卖决策：")
    for path in saved_files:
        print(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--watchlist", default=None, help="昨晚筛选观察池CSV，例如 watchlists/latest_watchlist.csv")
    ap.add_argument("--holdings", default=None, help="持仓CSV，格式：代码,名称,成本,股数")
    ap.add_argument("--code", default=None, help="单只股票代码")
    ap.add_argument("--cost", type=float, default=None, help="单只股票成本")
    ap.add_argument("--shares", type=int, default=None, help="单只股票股数")
    ap.add_argument("--only-action", action="store_true", help="只显示有明确动作的股票")
    ap.add_argument("--show-checks", action="store_true", help="显示买入确认检查项")
    ap.add_argument("--minute-period", default="5", choices=["1", "5", "15", "30", "60"], help="分钟线周期")
    ap.add_argument("--capital", type=float, default=None, help="账户总资金；填写后自动按止损距离估算建议股数")
    ap.add_argument("--risk-pct", type=float, default=1.0, help="单笔账户风险百分比，默认 1%%")
    ap.add_argument("--sleep", type=float, default=0.1)
    ap.add_argument("--workers", type=int, default=4, help="并行分析股票数，默认 4")
    args = ap.parse_args()

    if not args.watchlist and not args.holdings and not args.code:
        default = Path("watchlists/latest_watchlist.csv")
        if default.exists():
            args.watchlist = str(default)
        else:
            print("请指定 --watchlist、--holdings 或 --code。")
            print("例如：python a_stock_live_decision_v8.py --watchlist watchlists/latest_watchlist.csv --holdings holdings.csv")
            return

    items = merge_watchlist_and_holdings(args.watchlist, args.holdings, args.code, args.cost, args.shares)
    if not items:
        print("没有可分析的股票。")
        return

    market_context = get_market_context()
    print(f"开始实时买卖一体分析，共 {len(items)} 只。")
    print(f"大盘环境：{market_context.get('市场环境', '缺失')} | {market_context.get('市场建议', '')}")
    rows_by_index = {}

    worker_count = max(1, min(args.workers, len(items)))
    with ProcessPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(
                get_live_decision,
                item,
                minute_period=args.minute_period,
                market_context=market_context,
                capital=args.capital,
                account_risk_pct=args.risk_pct,
            ): (index, item)
            for index, item in enumerate(items)
        }
        for future in as_completed(futures):
            index, item = futures[future]
            try:
                rows_by_index[index] = future.result()
            except KeyboardInterrupt:
                raise
            except Exception as reason:
                print(f"[跳过] {item.get('代码', '')} {item.get('名称', '')}: {reason}")
            time.sleep(args.sleep)

    rows = [rows_by_index[index] for index in sorted(rows_by_index)]
    for d in rows:
        clear_action = (
            d["买入判断"] == "可以买小仓"
            or "止损" in d["最终动作"]
            or "减仓" in d["最终动作"]
            or "止盈" in d["最终动作"]
            or "加仓" in d["最终动作"]
        )
        if not args.only_action or clear_action:
            print_decision(d, show_checks=args.show_checks)

    save_results(rows)


if __name__ == "__main__":
    main()
